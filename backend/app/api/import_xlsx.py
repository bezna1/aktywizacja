from datetime import date, datetime, timedelta
from tempfile import SpooledTemporaryFile
from typing import Any

from fastapi import APIRouter, Depends, File, UploadFile
from openpyxl import load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.api.deps import require_roles
from app.db.session import get_db
from app.models import Participant, ParticipantTraining, ProgramAssignment, ProgramType, Training, TrainingStatus, User, UserRole
from app.schemas import ImportResult
from app.services.audit import write_audit
from app.validation import birth_date_from_pesel, participant_warnings

router = APIRouter(prefix="/import", tags=["import"])


HEADER_MAP = {
    "nazwisko i imię": "full_name",
    "uwagi": "notes",
    "telefon kontaktowy": "phone",
    "ZUS/PUP": "pup_zus_status",
    "wykluczenie": "exclusion",
    "niepełnosprawność": "disability_status",
    "zdrowotne": "health_status",
    "sytuacja życiowa": "life_situation",
    "pesel": "pesel",
    "wiek": "age",
    "płeć": "gender",
    "wykształcenie": "education",
    "województwo": "voivodeship",
    "powiat": "county",
    "gmina": "commune",
    "miejscowość": "city",
    "kod pocztowy": "postal_code",
    "ulica": "address",
    "status osoby na rynku pracy w chwili przystąpienia do projektu": "labor_market_status",
}


def split_name(value: str | None) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    parts = str(value).strip().split()
    if len(parts) == 1:
        return parts[0], None
    return " ".join(parts[1:]), parts[0]


@router.post("/xlsx", response_model=ImportResult)
async def import_xlsx(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(UserRole.admin, UserRole.worker)),
) -> ImportResult:
    errors: list[dict] = []
    imported = 0
    skipped = 0

    tmp = SpooledTemporaryFile(max_size=10_000_000)
    tmp.write(await file.read())
    tmp.seek(0)
    workbook = load_workbook(tmp, read_only=True, data_only=True)
    sheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    mapped = {index: HEADER_MAP[name] for index, name in enumerate(headers) if name in HEADER_MAP}

    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        raw = {}
        for index, field in mapped.items():
            value = row[index] if index < len(row) else None
            raw[field] = str(value).strip() if value is not None else None
        if not any(raw.values()):
            skipped += 1
            continue
        first_name, last_name = split_name(raw.pop("full_name", None))
        data = {
            **raw,
            "first_name": first_name,
            "last_name": last_name,
            "birth_date": birth_date_from_pesel(raw.get("pesel")),
            "age": int(raw["age"]) if raw.get("age") and str(raw["age"]).isdigit() else None,
        }
        warnings = participant_warnings(data)
        participant = Participant(**data, is_complete=not warnings, warning_reasons=warnings)
        db.add(participant)
        try:
            db.flush()
            imported += 1
        except IntegrityError:
            db.rollback()
            skipped += 1
            errors.append({"row": row_number, "error": "Duplikat lub błędne dane unikalne"})

    write_audit(db, user, "import_xlsx", "participant", None, {"file": file.filename, "imported": imported, "at": datetime.utcnow().isoformat()})
    db.commit()
    return ImportResult(imported=imported, skipped=skipped, errors=errors[:100])


# ---------------------------------------------------------------------------
# Schedule import helpers
# ---------------------------------------------------------------------------

_EXCEL_EPOCH = datetime(1899, 12, 30)


def _parse_date(value: Any) -> date | None:
    """Convert Excel cell value to date — handles datetime, serial int, and DD.MM.YYYY strings."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (_EXCEL_EPOCH + timedelta(days=int(value))).date()
        except (OverflowError, ValueError):
            return None
    if isinstance(value, str):
        v = value.strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try:
                return datetime.strptime(v, fmt).date()
            except ValueError:
                continue
    return None


def _cell(row: tuple, col: int) -> Any:
    return row[col] if col < len(row) else None


def _fmt_meeting(d: date | None, time_str: str | None) -> str | None:
    parts = []
    if d:
        parts.append(d.strftime("%d.%m.%Y"))
    if time_str and str(time_str).strip():
        parts.append(str(time_str).strip())
    return " ".join(parts) if parts else None


def _build_schedule_notes(meetings: list[tuple[date | None, str | None]], labels: list[str] | None = None) -> str:
    lines = []
    for i, (d, t) in enumerate(meetings):
        label = labels[i] if labels and i < len(labels) else f"Spotkanie {i + 1}"
        entry = _fmt_meeting(d, t)
        if entry:
            lines.append(f"{label}: {entry}")
    return "\n".join(lines)


# col indices (0-based) for each program block
_DORADCA_ZAWODOWY_COLS = [(17, 18), (19, 20)]
_KOMPETENCJE_CYFROWE_COLS = [(21, 22)]
_PSYCHOLOG_COLS = [(23, 24), (25, 26)]
_DORADCA_PRAWNY_COLS = [(30, 31)]
_POSREDNICTWO_COLS = [(32, 33), (34, 35), (36, 37), (38, 39)]
_POSREDNICTWO_OWN_WORK_COL = 40
_WARSZTATY_SPOL_COLS = [(41, None), (42, None), (43, None), (44, None)]
_WARSZTATY_ZAW_COLS = [(45, None), (46, None)]
_SZKOLENIE_NAME_COL = 27
_SZKOLENIE_TERM_COL = 28
_SZKOLENIE_HOURS_COL = 29


def _upsert_assignment(db: Session, participant_id: int, program_type: ProgramType, planned_at: date | None, notes: str | None) -> None:
    existing = (
        db.query(ProgramAssignment)
        .filter_by(participant_id=participant_id, program_type=program_type)
        .first()
    )
    if existing:
        existing.planned_at = datetime.combine(planned_at, datetime.min.time()) if planned_at else None
        existing.notes = notes
    else:
        db.add(ProgramAssignment(
            participant_id=participant_id,
            program_type=program_type,
            status="zaplanowany",
            planned_at=datetime.combine(planned_at, datetime.min.time()) if planned_at else None,
            notes=notes,
        ))


def _upsert_vocational_training(db: Session, participant_id: int, name: str, term: str | None, hours: int) -> None:
    training = db.query(Training).filter_by(name=name).first()
    if not training:
        training = Training(
            name=name,
            schedule=term,
            hours_count=hours,
            status=TrainingStatus.active,
        )
        db.add(training)
        db.flush()

    existing_pt = db.query(ParticipantTraining).filter_by(participant_id=participant_id, training_id=training.id).first()
    if not existing_pt:
        db.add(ParticipantTraining(
            participant_id=participant_id,
            training_id=training.id,
            notes=term,
        ))


def _import_schedule_row(db: Session, row: tuple, participant_id: int) -> None:
    # DORADCA ZAWODOWY
    meetings = [(_parse_date(_cell(row, dc)), str(_cell(row, tc) or "")) for dc, tc in _DORADCA_ZAWODOWY_COLS]
    if any(d for d, _ in meetings):
        notes = _build_schedule_notes(meetings, ["I spotkanie", "II spotkanie"])
        _upsert_assignment(db, participant_id, ProgramType.career_advisor, meetings[0][0], notes)

    # KOMPETENCJE CYFROWE
    kc_date = _parse_date(_cell(row, 21))
    kc_time = str(_cell(row, 22) or "")
    if kc_date:
        notes = _build_schedule_notes([(kc_date, kc_time)], ["Szkolenie"])
        _upsert_assignment(db, participant_id, ProgramType.digital_skills, kc_date, notes)

    # PSYCHOLOG
    meetings = [(_parse_date(_cell(row, dc)), str(_cell(row, tc) or "")) for dc, tc in _PSYCHOLOG_COLS]
    if any(d for d, _ in meetings):
        notes = _build_schedule_notes(meetings, ["I spotkanie", "II spotkanie"])
        _upsert_assignment(db, participant_id, ProgramType.psychologist, meetings[0][0], notes)

    # SZKOLENIE ZAWODOWE
    sz_name = _cell(row, _SZKOLENIE_NAME_COL)
    sz_term = _cell(row, _SZKOLENIE_TERM_COL)
    sz_hours_raw = _cell(row, _SZKOLENIE_HOURS_COL)
    if sz_name:
        sz_hours = int(sz_hours_raw) if sz_hours_raw and str(sz_hours_raw).strip().isdigit() else (int(float(sz_hours_raw)) if sz_hours_raw else 0)
        _upsert_vocational_training(db, participant_id, str(sz_name).strip(), str(sz_term).strip() if sz_term else None, sz_hours)

    # DORADCA PRAWNY
    dp_date = _parse_date(_cell(row, 30))
    dp_time = str(_cell(row, 31) or "")
    if dp_date:
        notes = _build_schedule_notes([(dp_date, dp_time)], ["Wizyta"])
        _upsert_assignment(db, participant_id, ProgramType.legal_advisor, dp_date, notes)

    # POŚREDNICTWO ZAWODOWE
    meetings = [(_parse_date(_cell(row, dc)), str(_cell(row, tc) or "")) for dc, tc in _POSREDNICTWO_COLS]
    own_work = _cell(row, _POSREDNICTWO_OWN_WORK_COL)
    if any(d for d, _ in meetings):
        labels = ["I spotkanie", "II spotkanie", "III spotkanie", "IV spotkanie"]
        notes = _build_schedule_notes(meetings, labels)
        if own_work:
            notes += f"\nPraca własna: {own_work} h"
        _upsert_assignment(db, participant_id, ProgramType.job_placement, meetings[0][0], notes)

    # WARSZTATY AKTYWIZACJI SPOŁECZNEJ
    meetings_spol = [(_parse_date(_cell(row, dc)), None) for dc, _ in _WARSZTATY_SPOL_COLS]
    if any(d for d, _ in meetings_spol):
        notes = _build_schedule_notes(meetings_spol, ["I spotkanie", "II spotkanie", "III spotkanie", "IV spotkanie"])
        _upsert_assignment(db, participant_id, ProgramType.social_activation, meetings_spol[0][0], notes)

    # WARSZTATY AKTYWIZACJI ZAWODOWEJ
    meetings_zaw = [(_parse_date(_cell(row, dc)), None) for dc, _ in _WARSZTATY_ZAW_COLS]
    if any(d for d, _ in meetings_zaw):
        notes = _build_schedule_notes(meetings_zaw, ["I spotkanie", "II spotkanie"])
        _upsert_assignment(db, participant_id, ProgramType.work_activation, meetings_zaw[0][0], notes)


@router.post("/schedule", response_model=ImportResult)
async def import_schedule(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(require_roles(UserRole.admin, UserRole.worker)),
) -> ImportResult:
    """Import harmonogramu szkoleń z pliku XLSX (wiersze 3+).

    Dopasowuje uczestników po PESEL. Tworzy ProgramAssignment dla każdego
    rodzaju wsparcia oraz Training + ParticipantTraining dla szkoleń zawodowych.
    """
    errors: list[dict] = []
    imported = 0
    skipped = 0

    tmp = SpooledTemporaryFile(max_size=10_000_000)
    tmp.write(await file.read())
    tmp.seek(0)
    workbook = load_workbook(tmp, read_only=False, data_only=True)
    sheet = workbook.active

    pesel_col: int | None = None
    for idx, cell in enumerate(next(sheet.iter_rows(min_row=1, max_row=1))):
        if cell.value and str(cell.value).strip().lower() == "pesel":
            pesel_col = idx
            break

    if pesel_col is None:
        return ImportResult(imported=0, skipped=0, errors=[{"row": 1, "error": "Nie znaleziono kolumny PESEL"}])

    for row_number, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        pesel_raw = row[pesel_col] if pesel_col < len(row) else None
        if not pesel_raw:
            skipped += 1
            continue
        pesel = str(pesel_raw).strip()

        participant = db.query(Participant).filter_by(pesel=pesel).first()
        if not participant:
            skipped += 1
            errors.append({"row": row_number, "error": f"Uczestnik z PESEL {pesel} nie istnieje w bazie"})
            continue

        try:
            _import_schedule_row(db, row, participant.id)
            db.flush()
            imported += 1
        except Exception as exc:
            db.rollback()
            skipped += 1
            errors.append({"row": row_number, "error": str(exc)})

    write_audit(db, user, "import_schedule", "participant", None, {"file": file.filename, "imported": imported, "at": datetime.utcnow().isoformat()})
    db.commit()
    return ImportResult(imported=imported, skipped=skipped, errors=errors[:100])
